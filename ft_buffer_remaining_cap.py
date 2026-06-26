import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter, FilterColumn

# ==================================================
# CONFIGURACIÓN GENERAL
# ==================================================

# Ruta del Excel final
RUTA_EXCEL = Path(__file__).parent / "FT buffer remaining cap.xlsx"
os.makedirs(RUTA_EXCEL.parent, exist_ok=True)

# Inputs del scraping
TICKERS_F = ["FJAN","FFEB","FMAR","FAPR","FMAY","FJUN","FJUL","FAUG","FSEP","FOCT","FNOV","FDEC"]
TICKERS_G = ["GJAN","GFEB","GMAR","GAPR","GMAY","GJUN","GJUL","GAUG","GSEP","GOCT","GNOV","GDEC"]
TICKERS = TICKERS_F + TICKERS_G

# Métricas que van al resumen (en el orden deseado del reporte)
TITULOS_OBJETIVO = [
    "Remaining Buffer (Net)",
    "Remaining Cap (Net)",
    "Fund Value/Return",
    "Reference Asset Value/Return",
    "Reference Asset Return to Realize the Cap",
    "Downside Before Buffer (Net)"
]

# ==================================================
# UTILIDADES (LIMPIEZA DE DATOS)
# ==================================================

def valor_prioritario(x):
    if pd.isna(x):
        return None
    x = str(x)
    m = re.search(r"\(([^)]+)\)", x)
    if m:
        return m.group(1).strip()
    if "/" in x:
        return x.split("/")[-1].strip()
    return x.strip()

# ==================================================
# SCRAPING (SI CAMBIAS DE WEB, TOCAS AQUÍ)
# ==================================================

# Extrae la fecha "As of" y corrige MM/DD/YYYY → DD/MM/YYYY
def extraer_fecha_asof(soup):
    bloque = soup.find("div", class_="ftmas CEFFieldLabel")
    if not bloque:
        return None
    texto = bloque.get_text(" ", strip=True)
    m = re.search(r"as of\s+([0-9/]+)", texto, re.I)
    if not m:
        return None
    # La web entrega MM/DD/YYYY → invertir a DD/MM/YYYY
    partes = m.group(1).split("/")
    if len(partes) == 3:
        fecha_corregida = f"{partes[1]}/{partes[0]}/{partes[2]}"
        return f"As of {fecha_corregida}"
    return f"As of {m.group(1)}"


# Scraping de un solo ticker
def scrapear_ticker(ticker):

    url = f"https://www.ftportfolios.com/Retail/Etf/EtfSummary.aspx?Ticker={ticker}"
    r = requests.get(url, timeout=15)
    if r.status_code != 200:
        return None, []

    soup = BeautifulSoup(r.content, "html.parser")

    # Fecha (metadata)
    fecha = extraer_fecha_asof(soup)

    # Localización de la tabla principal
    filtro_Tabla = soup.find("div", class_="ftmas CEFFieldLabel")
    if filtro_Tabla is None:
        return fecha, []

    # Filas de datos
    datos = filtro_Tabla.find_all(
        "div", class_="flex justify-content-between align-items-end"
    )

    # Encabezados
    titulos = filtro_Tabla.find_all(
        "div", class_="silverBox fundControlSeperatorBar CEFFieldLabel ps-1 bold"
    )

    resultados = []

    # Encabezado combinado
    if len(titulos) >= 2:
        resultados.append({
            "Ticker": ticker,
            "Titulo": titulos[0].get_text(strip=True),
            "Valor": titulos[1].get_text(strip=True)
        })

    # Filas normales
    for dato in datos:
        valores = dato.find_all("div", recursive=False)
        if len(valores) == 2:
            resultados.append({
                "Ticker": ticker,
                "Titulo": valores[0].get_text(strip=True),
                "Valor": valores[1].get_text(strip=True)
            })

    return fecha, resultados

# ==================================================
# TRANSFORMACIÓN DE DATOS (PANDAS)
# ==================================================

def construir_dataframe(resultados):
    df = pd.DataFrame(resultados)
    df = df[df["Titulo"].notna() & (df["Titulo"].str.strip() != "")]
    orden = df["Titulo"].drop_duplicates().tolist()
    df["Titulo"] = pd.Categorical(df["Titulo"], orden, ordered=True)
    df = df.pivot(index="Titulo", columns="Ticker", values="Valor").reset_index()
    df.columns.name = None
    return df


def construir_resumen(df):
    df_res = df.copy()
    for col in df_res.columns[1:]:
        df_res[col] = df_res[col].apply(valor_prioritario)

    df_res = df_res[df_res["Titulo"].isin(TITULOS_OBJETIVO)]
    df_long = df_res.melt("Titulo", var_name="Ticker", value_name="Valor")

    pivot = df_long.pivot(
        index="Ticker",
        columns="Titulo",
        values="Valor"
    ).reset_index()

    # Reordenar columnas según TITULOS_OBJETIVO
    orden_cols = ["Ticker"] + [t for t in TITULOS_OBJETIVO if t in pivot.columns]
    pivot = pivot[orden_cols]

    # Renombrar columnas para el reporte
    pivot = pivot.rename(columns={
        "Fund Value/Return": "Fund Return",
        "Reference Asset Value/Return": "Reference Asset Return",
        "Reference Asset Return to Realize the Cap": "Reference Asset Return to Realize Cap"
    })

    # Ordenar filas cronológicamente (JAN→DEC)
    orden_tickers = TICKERS_F + TICKERS_G
    pivot["Ticker"] = pd.Categorical(pivot["Ticker"], orden_tickers, ordered=True)
    return pivot.sort_values("Ticker").reset_index(drop=True)

# ==================================================
# EXCEL / FORMATO (SI CAMBIAS REPORTE, TOCAS AQUÍ)
# ==================================================

def tabla_formato(ws, start_row, titulo, df_tabla, numero_tabla):
    ws.cell(row=start_row, column=1).value = titulo
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    end_row = start_row + len(df_tabla) + 1
    end_col = len(df_tabla.columns)

    ref = f"{ws.cell(start_row + 1, 1).coordinate}:{ws.cell(end_row, end_col).coordinate}"

    tabla = Table(displayName=f"Tabla_{numero_tabla}", ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    # Mantener la tabla de Excel, pero ocultar los botones de filtro.
    filtro = AutoFilter(ref=ref)
    filtro.filterColumn = [FilterColumn(colId=i, showButton=False) for i in range(end_col)]
    tabla.autoFilter = filtro

    ws.add_table(tabla)

    for r in range(start_row + 2, end_row + 1):
        for c in range(2, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
                if cell.value < 0:
                    cell.font = Font(color="FF0000")

def escribir_excel(df_base, df_f, df_g):
    with pd.ExcelWriter(RUTA_EXCEL, engine="openpyxl", mode="w") as writer:
        df_base.to_excel(writer, sheet_name="Datos ETFs", index=False)
        df_f.to_excel(writer, sheet_name="Resumen F", index=False)
        df_g.to_excel(writer, sheet_name="Resumen G", index=False)


def aplicar_formato_excel(fecha, df_base, df_f, df_g):
    wb = load_workbook(RUTA_EXCEL)

    configuracion_hojas = [
        ("Datos ETFs", "Datos ETFs", df_base, 1, False),
        ("Resumen F", "Resumen F", df_f, 2, True),
        ("Resumen G", "Resumen G", df_g, 3, True),
    ]

    for hoja, titulo, df_tabla, numero_tabla, incluir_fecha in configuracion_hojas:
        ws = wb[hoja]
        if incluir_fecha:
            ws.insert_rows(1, 2)
            ws["A1"] = fecha
            ws["A1"].font = Font(bold=True)
            ws["A1"].alignment = Alignment(horizontal="left")
            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=ws.max_column
            )
            tabla_formato(ws, start_row=2, titulo=titulo, df_tabla=df_tabla, numero_tabla=numero_tabla)
        else:
            ws.insert_rows(1, 1)
            tabla_formato(ws, start_row=1, titulo=titulo, df_tabla=df_tabla, numero_tabla=numero_tabla)

    wb.save(RUTA_EXCEL)

# ==================================================
# MAIN (ORQUESTADOR)
# ==================================================

def main():
    resultados = []
    fecha_global = None

    for ticker in TICKERS:
        print(f"Procesando {ticker}...")
        fecha, datos = scrapear_ticker(ticker)
        resultados.extend(datos)
        if fecha_global is None and fecha:
            fecha_global = fecha

    df_base = construir_dataframe(resultados)
    df_resumen = construir_resumen(df_base)

    df_f = df_resumen[df_resumen["Ticker"].isin(TICKERS_F)]
    df_g = df_resumen[df_resumen["Ticker"].isin(TICKERS_G)]

    escribir_excel(df_base, df_f, df_g)
    aplicar_formato_excel(fecha_global, df_base, df_f, df_g)

    print("Proceso terminado | Fecha:", fecha_global)


if __name__ == "__main__":
    main()
